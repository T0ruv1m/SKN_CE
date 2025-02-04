import os
import time
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

def transform_to_table(file_path):

    try:
        wb = load_workbook(file_path)
        sheet = wb.active

        min_row = sheet.min_row
        max_row = sheet.max_row
        min_col = sheet.min_column
        max_col = sheet.max_column
        table_range = f"{sheet.cell(row=min_row, column=min_col).coordinate}:{sheet.cell(row=max_row,column=max_col).coordinate}"

        table = Table(displayName="Datatable", ref=table_range)

        style = TableStyleInfo(
                    name="TableStyleMedium9",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=True,
                )
        table.tableStyleInfo = style
        sheet.add_table(table)

        wb.save(file_path)
        print(f"Tabela Criada com Sucesso em {file_path}")
    except Exception as e:
        print(f"Erro ao criar tabela!")